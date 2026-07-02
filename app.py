"""
ShipScan v7  (accuracy + precision update)

  Tab 1 — Targeted Extract: filter by filename / subject / in-PDF text,
          preview before downloading, auto-download matches, log to Excel.
  Tab 2 — Browse & Pick: list ALL PDFs from a sender newest→oldest,
          checkbox-select, save chosen.

WHAT CHANGED vs v6 — why you now get the RIGHT files
  v6 only matched on the *exact* attachment filename, which silently missed
  files whenever the bill number lived in the subject or inside the PDF, and
  it could overwrite / skip distinct files that happened to share a name.

  v7 adds:
    • Match modes — Contains (default, intuitive) / Exact / Wildcard / Regex.
        "1234567" now finds "SB_1234567_2024.pdf" without needing *…*.
    • Match the email SUBJECT too (opt-in) — catches generically-named PDFs.
    • Search INSIDE the PDF text (opt-in, slower) — finds the bill by a number
        printed on the document even when the filename is "scan.pdf".
    • Preview / dry-run — see exactly what WOULD download before committing.
    • Safe de-duplication — two different files named "invoice.pdf" no longer
        clobber each other; a genuine re-download of the same file still skips.
    • PDF validation — a file is only marked "Saved" if it really is a PDF
        (%PDF header + sane size), so silent corruption shows as "Failed".
    • Broader attachment detection — application/x-pdf, octet-stream/*.pdf,
        and PDFs sent without a filename are no longer missed.

  The fast metadata-first scan from v6 is preserved: bodies are only fetched
  for parts that are actually candidates (and, with text search off, only for
  ones that already matched on name/subject).

Dependencies:  flask  openpyxl  imapclient  pypdf(optional, for in-PDF search)
"""

import os
import io
import re
import base64
import quopri
import fnmatch
import threading
from datetime import datetime, timedelta
from email.header import decode_header

from flask import Flask, render_template, request, jsonify, send_file
from imapclient import IMAPClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# pypdf is optional — the app runs fine without it; in-PDF text search is
# simply disabled (and reported) when it's missing.
try:
    from pypdf import PdfReader
    HAS_PYPDF = True
except Exception:
    HAS_PYPDF = False

app = Flask(__name__)

IMAP_HOST = "imap.gmail.com"
BATCH = 300            # messages per metadata round-trip
PDF_MAGIC = b"%PDF"
MIN_PDF_BYTES = 800    # anything smaller is almost certainly truncated/garbage
TEXT_SCAN_PAGES = 12   # cap pages read when searching inside a PDF

# ── Job state (targeted extract) ─────────────────────────────────────────────
job = {"running": False, "done": False, "logs": [], "rows": [], "summary": {}, "excel_path": ""}

# ── Browse state ──────────────────────────────────────────────────────────────
browse = {"running": False, "done": False, "pdfs": [], "error": ""}

# one lock so a second tab/click can't trample an in-flight run
_lock = threading.Lock()


# ── Helpers ───────────────────────────────────────────────────────────────────

def decode_str(val):
    if not val:
        return ""
    if isinstance(val, (bytes, bytearray)):
        val = val.decode("utf-8", errors="replace")
    out = []
    for part, cs in decode_header(val):
        if isinstance(part, bytes):
            out.append(part.decode(cs or "utf-8", errors="replace"))
        else:
            out.append(part)
    return "".join(out)


def make_server(gmail_addr, app_pass):
    server = IMAPClient(IMAP_HOST, ssl=True)
    server.login(gmail_addr, app_pass)
    return server


def matches_filter(name, filters, mode="contains"):
    """
    Match a NAME (filename or subject line) against the user's filters.

    mode:
      contains  — substring match (default; what people intuitively expect).
                  Explicit * / ? in a term still switch that term to wildcard.
      exact     — whole-name equality, ignoring the extension.
      wildcard  — fnmatch; a term without * / ? is auto-wrapped as *term*.
      regex     — Python re.search; a broken pattern degrades to a literal
                  substring match instead of throwing.
    Returns (matched: bool, which_filter: str).
    """
    a = (name or "").lower().strip()
    for f in filters:
        f = f.strip()
        if not f:
            continue
        fl = f.lower()
        try:
            if mode == "regex":
                if re.search(fl, a):
                    return True, f
            elif mode == "wildcard":
                pat = fl if ("*" in fl or "?" in fl) else f"*{fl}*"
                if fnmatch.fnmatch(a, pat):
                    return True, f
            elif mode == "exact":
                if a == fl or os.path.splitext(a)[0] == os.path.splitext(fl)[0]:
                    return True, f
            else:  # contains — but still honour explicit wildcards
                if "*" in fl or "?" in fl:
                    if fnmatch.fnmatch(a, fl):
                        return True, f
                elif fl in a:
                    return True, f
        except re.error:
            if fl in a:          # bad regex → behave like contains
                return True, f
    return False, ""


def text_matches(text, filters, mode="contains"):
    """Match against text extracted from inside a PDF."""
    t = (text or "").lower()
    if not t:
        return False, ""
    for f in filters:
        f = f.strip()
        if not f:
            continue
        fl = f.lower()
        if mode == "regex":
            try:
                if re.search(fl, t):
                    return True, f
            except re.error:
                if fl in t:
                    return True, f
        else:
            needle = fl.strip("*?")
            if needle and needle in t:
                return True, f
    return False, ""


def log(msg, level="info"):
    job["logs"].append({"t": datetime.now().strftime("%H:%M:%S"), "msg": msg, "lvl": level})


def _dec(x):
    return x.decode() if isinstance(x, (bytes, bytearray)) else (x or "")


def _walk_bodystructure(part, prefix=""):
    """
    Walk an imapclient BODYSTRUCTURE node WITHOUT downloading any bodies.
    Yields dicts: {part, mediatype, filename, enc_size, encoding} for every leaf part.
    """
    out = []
    if isinstance(part[0], (list, tuple)):
        for i, child in enumerate(part[0], 1):
            num = f"{i}" if not prefix else f"{prefix}.{i}"
            out += _walk_bodystructure(child, num)
        return out

    num = prefix or "1"
    mediatype = (_dec(part[0]).lower() + "/" + _dec(part[1]).lower())
    encoding = _dec(part[5]).lower()
    enc_size = part[6] if len(part) > 6 and isinstance(part[6], int) else 0

    filename = None
    disp = part[8] if len(part) > 8 else None
    if isinstance(disp, (list, tuple)) and len(disp) > 1 and isinstance(disp[1], (list, tuple)):
        d = disp[1]
        for k in range(0, len(d) - 1, 2):
            if _dec(d[k]).lower() == "filename":
                filename = _dec(d[k + 1])
    if not filename and isinstance(part[2], (list, tuple)):
        ct = part[2]
        for k in range(0, len(ct) - 1, 2):
            if _dec(ct[k]).lower() == "name":
                filename = _dec(ct[k + 1])

    if filename:
        filename = decode_str(filename)

    out.append({
        "part": num,
        "mediatype": mediatype,
        "filename": filename,
        "enc_size": enc_size,
        "encoding": encoding,
    })
    return out


# ── Attachment type groups (what the UI's "Files" chips map to) ───────────────
# Nothing here opens a file — these are just filename-extension buckets.
TYPE_GROUPS = {
    "pdf":     {".pdf"},
    "excel":   {".xlsx", ".xlsm", ".xls", ".csv", ".tsv"},
    "word":    {".docx", ".doc", ".rtf", ".odt"},
    "image":   {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tif", ".tiff", ".heic"},
    "archive": {".zip", ".rar", ".7z", ".tar", ".gz", ".tgz"},
    "ppt":     {".pptx", ".ppt", ".odp"},
}

# Fallback when an attachment arrives with NO filename: guess an extension
# from its media type so it can still be matched/saved sensibly.
_MT_EXT = {
    "application/pdf": ".pdf", "application/x-pdf": ".pdf",
    "application/acrobat": ".pdf", "text/pdf": ".pdf", "text/x-pdf": ".pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/vnd.ms-excel": ".xls", "text/csv": ".csv",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/msword": ".doc", "application/rtf": ".rtf", "text/rtf": ".rtf",
    "image/jpeg": ".jpg", "image/png": ".png", "image/gif": ".gif",
    "image/webp": ".webp", "image/tiff": ".tif", "image/bmp": ".bmp",
    "application/zip": ".zip", "application/x-zip-compressed": ".zip",
    "application/x-rar-compressed": ".rar", "application/x-7z-compressed": ".7z",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": ".pptx",
    "application/vnd.ms-powerpoint": ".ppt",
}


def exts_for(groups):
    """
    Turn a list of UI groups (e.g. ['pdf','excel']) into a set of allowed
    extensions. ['all'] (or empty) returns None, meaning "accept every type".
    """
    if not groups or "all" in groups:
        return None
    exts = set()
    for g in groups:
        exts |= TYPE_GROUPS.get(g, set())
    return exts or None


def _attachment_parts(bodystructure, allowed_exts):
    """
    Return every *attachment* leaf part (a MIME part carrying a filename),
    optionally restricted to allowed_exts. allowed_exts=None → accept all.

    We never open or read a part here — only its MIME headers are inspected.
    Inline message bodies (text/plain, text/html with no filename) are skipped
    because real attachments carry a filename or a known attachment media type.
    """
    res = []
    auto = 0
    for p in _walk_bodystructure(bodystructure):
        fn = (p["filename"] or "").strip()
        mt = p["mediatype"]

        # recover a missing filename from the media type (e.g. nameless PDF)
        if not fn:
            guessed = _MT_EXT.get(mt)
            if not guessed:
                continue                       # nameless + unknown → it's a body
            auto += 1
            fn = f"attachment_{auto}{guessed}"
            p = dict(p, filename=fn)

        ext = os.path.splitext(fn)[1].lower()
        if allowed_exts is None or ext in allowed_exts:
            res.append(p)
    return res


def _decode_part_body(raw, encoding):
    if raw is None:
        return b""
    if encoding == "base64":
        try:
            return base64.decodebytes(raw if isinstance(raw, bytes) else raw.encode())
        except Exception:
            return base64.b64decode(re.sub(rb"\s", b"", raw))
    if encoding in ("quoted-printable", "quopri"):
        return quopri.decodestring(raw if isinstance(raw, bytes) else raw.encode())
    return raw if isinstance(raw, bytes) else raw.encode()


def _decoded_kb(enc_size, encoding):
    real = enc_size * 0.75 if encoding == "base64" else enc_size
    return round(real / 1024, 1)


def _expected_bytes(enc_size, encoding):
    return int(enc_size * 0.75) if encoding == "base64" else int(enc_size or 0)


def _envelope_date(env):
    try:
        if env and env.date:
            return env.date.strftime("%Y-%m-%d %H:%M"), env.date.timestamp()
    except Exception:
        pass
    return "", 0.0


def _safe_name(filename):
    return re.sub(r'[<>:"/\\|?*]', "_", filename or "").strip() or "attachment.pdf"


def _resolve_save_path(folder, filename, expected):
    """
    Decide where to write a download, avoiding silent loss.

      • If a same-named file already exists AND its size matches what we expect,
        treat it as the same document already on disk → ("exists", path).
      • If a same-named file exists but the size differs, it's a DIFFERENT
        document → write under name_1.pdf, name_2.pdf… → ("new", unique_path).
      • Otherwise just use the plain path → ("new", path).
    """
    safe = _safe_name(filename)
    path = os.path.join(folder, safe)
    if os.path.exists(path):
        try:
            on_disk = os.path.getsize(path)
        except OSError:
            on_disk = -1
        tol = max(2048, int(expected * 0.05)) if expected else 2048
        if expected and abs(on_disk - expected) <= tol:
            return "exists", path
        base, ext = os.path.splitext(safe)
        n = 1
        cand = os.path.join(folder, f"{base}_{n}{ext}")
        while os.path.exists(cand):
            n += 1
            cand = os.path.join(folder, f"{base}_{n}{ext}")
        return "new", cand
    return "new", path


def _validate_pdf(data):
    if not data:
        raise ValueError("empty part")
    if len(data) < MIN_PDF_BYTES:
        raise ValueError("file too small — likely truncated")
    head = data[:1024]
    if PDF_MAGIC not in head:
        raise ValueError("not a valid PDF (no %PDF header)")


def _validate_download(data, filename):
    """
    Sanity-check a downloaded attachment before it's marked Saved.
    PDFs get the strict %PDF check; every other type just has to be
    non-empty (we don't read or parse non-PDFs).
    """
    if not data:
        raise ValueError("empty part")
    ext = os.path.splitext(filename or "")[1].lower()
    if ext == ".pdf":
        _validate_pdf(data)
    elif len(data) < 8:
        raise ValueError("file too small / empty")


def _pdf_text(data):
    if not HAS_PYPDF:
        return ""
    try:
        reader = PdfReader(io.BytesIO(data))
        pages = reader.pages[:TEXT_SCAN_PAGES]
        return "\n".join((pg.extract_text() or "") for pg in pages)
    except Exception:
        return ""


def _search_criteria(sender_email, days_back, subjects=None):
    crit = ["FROM", sender_email]
    if days_back and days_back > 0:
        since = (datetime.now() - timedelta(days=days_back)).date()
        crit += ["SINCE", since]
    subs = [s.strip() for s in (subjects or []) if s and s.strip()]
    if len(subs) == 1:
        crit += ["SUBJECT", subs[0]]
    elif len(subs) > 1:
        node = ["SUBJECT", subs[-1]]
        for s in reversed(subs[:-1]):
            node = ["OR", ["SUBJECT", s], node]
        crit += node
    return crit


def _parse_subjects(config):
    raw = (config.get("subject", "") or "").strip()
    return [l.strip() for l in raw.splitlines() if l.strip()]


def _truthy(v):
    return str(v).strip().lower() in ("1", "true", "yes", "on") if not isinstance(v, bool) else v


# ── MODE 1: Targeted extract ──────────────────────────────────────────────────

def run_targeted(config):
    job.update({"running": True, "done": False, "logs": [], "rows": [], "summary": {}})

    gmail_addr   = config["gmail_address"].strip()
    app_pass     = config["app_password"].strip()
    sender_email = config["sender_email"].strip()
    days_back    = int(config.get("days_back", 30) or 0)
    save_folder  = config["save_folder"].strip()
    raw_filters  = config.get("filters", "").strip()
    filters      = [l.strip() for l in raw_filters.splitlines() if l.strip()]
    fetch_all    = len(filters) == 0

    match_mode   = (config.get("match_mode") or "contains").strip().lower()
    match_subj   = _truthy(config.get("match_subject"))
    search_text  = _truthy(config.get("search_text"))
    preview      = _truthy(config.get("preview"))

    file_types   = config.get("file_types") or ["pdf"]
    allowed_exts = exts_for(file_types)          # None == every type

    if search_text and not HAS_PYPDF:
        search_text = False
        log("In-PDF text search requested but pypdf isn't installed — skipping it. "
            "Run: pip install pypdf", "warn")

    subj_mode = "regex" if match_mode == "regex" else "contains"

    excel_path = os.path.join(save_folder, f"shipscan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    job["excel_path"] = excel_path

    types_label = "all types" if allowed_exts is None else "/".join(file_types)
    bits = []
    bits.append(f"all {types_label}" if fetch_all else f"{len(filters)} filter(s) · mode={match_mode}")
    if not fetch_all: bits.append(types_label)
    if match_subj:  bits.append("subject")
    if search_text: bits.append("in-PDF text")
    if preview:     bits.append("PREVIEW (no download)")
    log("Mode: " + " · ".join(bits))
    log("Connecting…")

    try:
        if not preview:
            os.makedirs(save_folder, exist_ok=True)
        server = make_server(gmail_addr, app_pass)
        log("Connected", "success")
    except Exception as e:
        log(f"Login failed: {e}", "error")
        log("Enable IMAP and use an App Password.", "warn")
        job.update({"running": False, "done": True})
        return

    log(f"Searching from: {sender_email}")
    subjects = _parse_subjects(config)
    if subjects:
        log(f"Subject pre-filter: {', '.join(subjects)}")
    try:
        server.select_folder("INBOX", readonly=True)
        uids = server.search(_search_criteria(sender_email, days_back, subjects))
        uids = sorted(uids, reverse=True)
        log(f"{len(uids)} email(s) found", "success" if uids else "warn")
    except Exception as e:
        log(f"Search failed: {e}", "error")
        try: server.logout()
        except Exception: pass
        job.update({"running": False, "done": True})
        return

    # ── Phase 1: scan metadata only ──────────────────────────────────────────
    # Each candidate carries how it matched (or that it still needs a text check)
    matches = []
    scanned = 0
    try:
        for start in range(0, len(uids), BATCH):
            chunk = uids[start:start + BATCH]
            meta = server.fetch(chunk, ["BODYSTRUCTURE", "ENVELOPE"])
            for uid in chunk:
                data = meta.get(uid)
                if not data:
                    continue
                bs = data.get(b"BODYSTRUCTURE")
                env = data.get(b"ENVELOPE")
                if bs is None:
                    continue
                subject = decode_str(env.subject) if (env and env.subject) else ""
                email_date, sort_key = _envelope_date(env)
                for p in _attachment_parts(bs, allowed_exts):
                    fn = p["filename"]
                    is_pdf = os.path.splitext(fn)[1].lower() == ".pdf"
                    reason = ""
                    needs_text = False

                    if fetch_all:
                        reason = "all"
                    else:
                        ok, mf = matches_filter(fn, filters, match_mode)
                        if ok:
                            reason = f"name:{mf}"
                        elif match_subj:
                            ok, mf = matches_filter(subject, filters, subj_mode)
                            if ok:
                                reason = f"subject:{mf}"
                        if not reason:
                            # text search only applies to PDFs; other types are
                            # never opened, so a non-matching name/subject drops them
                            if search_text and is_pdf:
                                needs_text = True          # decide after download
                            else:
                                continue

                    matches.append({
                        "uid": uid, "part": p["part"], "encoding": p["encoding"],
                        "filename": fn, "match_reason": reason, "needs_text": needs_text,
                        "subject": subject, "email_date": email_date, "sort_key": sort_key,
                        "size_kb": _decoded_kb(p["enc_size"], p["encoding"]),
                        "expected": _expected_bytes(p["enc_size"], p["encoding"]),
                    })
            scanned += len(chunk)
            log(f"Scanned {scanned}/{len(uids)} emails… ({len(matches)} candidate(s))")
    except Exception as e:
        log(f"Scan error: {e}", "error")

    confirmed = sum(1 for m in matches if not m["needs_text"])
    pending   = sum(1 for m in matches if m["needs_text"])
    msg = f"{confirmed} match(es)"
    if pending:
        msg += f" + {pending} to text-check"
    log(msg, "success" if matches else "warn")

    # ── Phase 2: per-email fetch (download / preview) ─────────────────────────
    rows = []
    by_uid = {}
    for m in matches:
        by_uid.setdefault(m["uid"], []).append(m)

    def add_row(m, status, saved_path="", reason=None):
        rows.append({
            "filename": m["filename"],
            "matched_filter": reason if reason is not None else m["match_reason"],
            "email_date": m["email_date"], "subject": m["subject"],
            "saved_path": saved_path, "status": status,
            "size_kb": m["size_kb"],
        })
        job["rows"] = list(rows)

    for uid in sorted(by_uid.keys(), reverse=True):
        items = by_uid[uid]

        # ---- PREVIEW: never download; just report what would happen ----------
        if preview:
            for m in items:
                if m["needs_text"]:
                    add_row(m, "Would Check", reason="needs text scan")
                else:
                    add_row(m, "Would Save", reason=m["match_reason"])
                    log(f"• Would save: {m['filename']}  ({m['match_reason']})")
            continue

        # ---- decide which parts we actually need to fetch --------------------
        need = []
        for m in items:
            if m["needs_text"]:
                need.append(m)              # must download to inspect text
                continue
            kind, path = _resolve_save_path(save_folder, m["filename"], m["expected"])
            m["save_path"] = path
            if kind == "exists":
                add_row(m, "Already Exists", path)
                log(f"⏭ Exists: {m['filename']}", "warn")
            else:
                need.append(m)
        if not need:
            continue

        try:
            keys = [f"BODY.PEEK[{m['part']}]" for m in need]
            resp = server.fetch([uid], keys)
            item = resp.get(uid, {})
            for m in need:
                want = f"BODY[{m['part']}]".encode()
                raw = item.get(want)
                if raw is None:
                    for k, v in item.items():
                        if k.startswith(b"BODY[") and k.endswith(b"]") and v:
                            raw = v
                            break
                try:
                    pdf_data = _decode_part_body(raw, m["encoding"])

                    # in-PDF text check for files that didn't match name/subject
                    if m["needs_text"]:
                        ok, mf = text_matches(_pdf_text(pdf_data), filters, match_mode)
                        if not ok:
                            add_row(m, "No Match", reason="text: none")
                            log(f"· Skipped (no text match): {m['filename']}")
                            continue
                        m["match_reason"] = f"text:{mf}"
                        kind, path = _resolve_save_path(save_folder, m["filename"], m["expected"])
                        m["save_path"] = path
                        if kind == "exists":
                            add_row(m, "Already Exists", path, reason=m["match_reason"])
                            log(f"⏭ Exists: {m['filename']}", "warn")
                            continue

                    _validate_download(pdf_data, m["filename"])
                    with open(m["save_path"], "wb") as f:
                        f.write(pdf_data)
                    add_row(m, "Saved", m["save_path"])
                    log(f"✓ Saved: {os.path.basename(m['save_path'])}  ({m['match_reason']})", "success")
                except Exception as e:
                    add_row(m, "Failed")
                    log(f"✗ Failed: {m['filename']} — {e}", "error")
        except Exception as e:
            for m in need:
                add_row(m, "Failed")
            log(f"✗ Fetch failed for one email — {e}", "error")

    try: server.logout()
    except Exception: pass

    if not rows:
        log("No matching PDFs found.", "warn")

    saved   = sum(1 for r in rows if r["status"] == "Saved")
    failed  = sum(1 for r in rows if r["status"] == "Failed")
    exists  = sum(1 for r in rows if r["status"] == "Already Exists")
    skipped = sum(1 for r in rows if r["status"] in ("No Match", "Would Check"))
    would   = sum(1 for r in rows if r["status"] == "Would Save")
    job["summary"] = {"total": len(rows), "saved": saved, "failed": failed,
                      "exists": exists, "skipped": skipped, "would": would,
                      "preview": preview}

    if rows and not preview:
        try:
            _write_excel(rows, excel_path)
            log(f"Excel saved → {os.path.basename(excel_path)}", "success")
        except Exception as e:
            log(f"Excel error: {e}", "error")

    if preview:
        log(f"Preview done — {would} would save · {skipped} need a text scan", "success")
    else:
        log(f"Done — {saved} saved · {failed} failed · {exists} existed"
            + (f" · {skipped} skipped" if skipped else ""),
            "success" if not failed else "warn")
    job.update({"running": False, "done": True})


def _write_excel(rows, path):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ["#", "Filename", "Matched Via", "Email Date", "Subject",
               "Size (KB)", "Saved Path", "Status"]
    hfill = PatternFill("solid", start_color="4F46E5")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    thin  = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[1].height = 26
    fills = {
        "Saved":          PatternFill("solid", start_color="DCFCE7"),
        "Failed":         PatternFill("solid", start_color="FEE2E2"),
        "Already Exists": PatternFill("solid", start_color="EDE9FE"),
        "No Match":       PatternFill("solid", start_color="F1F5F9"),
    }
    dfont  = Font(name="Calibri", size=9)
    dalign = Alignment(vertical="center")
    for i, row in enumerate(rows):
        r = i + 2
        vals = [i + 1, row["filename"], row["matched_filter"], row["email_date"],
                row["subject"], row.get("size_kb", ""), row["saved_path"], row["status"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = dfont; cell.alignment = dalign; cell.border = thin
            if c == 8:
                cell.fill = fills.get(v, PatternFill())
    for c, w in enumerate([4, 30, 22, 16, 36, 10, 46, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # tiny summary sheet
    s = wb.create_sheet("Summary")
    s.column_dimensions["A"].width = 20
    s.column_dimensions["B"].width = 12
    stat = {}
    for row in rows:
        stat[row["status"]] = stat.get(row["status"], 0) + 1
    s.cell(row=1, column=1, value="Status").font = hfont
    s.cell(row=1, column=2, value="Count").font = hfont
    s.cell(row=1, column=1).fill = hfill
    s.cell(row=1, column=2).fill = hfill
    for i, (k, v) in enumerate(sorted(stat.items()), start=2):
        s.cell(row=i, column=1, value=k)
        s.cell(row=i, column=2, value=v)
    s.cell(row=len(stat) + 3, column=1, value="Generated")
    s.cell(row=len(stat) + 3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    wb.save(path)


# ── MODE 2: Browse all PDFs ───────────────────────────────────────────────────

def run_browse(config):
    browse.update({"running": True, "done": False, "pdfs": [], "error": ""})

    gmail_addr   = config["gmail_address"].strip()
    app_pass     = config["app_password"].strip()
    sender_email = config["sender_email"].strip()
    days_back    = int(config.get("days_back", 30) or 0)

    try:
        server = make_server(gmail_addr, app_pass)
    except Exception as e:
        browse.update({"running": False, "done": True, "error": f"Login failed: {e}"})
        return

    try:
        server.select_folder("INBOX", readonly=True)
        uids = server.search(_search_criteria(sender_email, days_back, _parse_subjects(config)))
        uids = sorted(uids, reverse=True)
    except Exception as e:
        browse.update({"running": False, "done": True, "error": f"Search failed: {e}"})
        try: server.logout()
        except Exception: pass
        return

    allowed_exts = exts_for(config.get("file_types") or ["pdf"])
    pdfs = []
    try:
        for start in range(0, len(uids), BATCH):
            chunk = uids[start:start + BATCH]
            meta = server.fetch(chunk, ["BODYSTRUCTURE", "ENVELOPE"])
            for uid in chunk:
                data = meta.get(uid)
                if not data:
                    continue
                bs = data.get(b"BODYSTRUCTURE")
                env = data.get(b"ENVELOPE")
                if bs is None:
                    continue
                subject = decode_str(env.subject) if (env and env.subject) else ""
                email_date, sort_key = _envelope_date(env)
                for p in _attachment_parts(bs, allowed_exts):
                    pdfs.append({
                        "id":         f"{uid}_{p['part']}",
                        "msg_id":     str(uid),
                        "uid":        uid,
                        "part":       p["part"],
                        "part_index": p["part"],
                        "encoding":   p["encoding"],
                        "filename":   p["filename"],
                        "email_date": email_date,
                        "sort_key":   sort_key,
                        "subject":    subject,
                        "size_kb":    _decoded_kb(p["enc_size"], p["encoding"]),
                        "expected":   _expected_bytes(p["enc_size"], p["encoding"]),
                    })
    except Exception as e:
        browse.update({"running": False, "done": True, "error": f"Scan failed: {e}"})
        try: server.logout()
        except Exception: pass
        return

    try: server.logout()
    except Exception: pass

    pdfs.sort(key=lambda x: x["sort_key"], reverse=True)
    browse.update({"running": False, "done": True, "pdfs": pdfs, "error": ""})


def save_selected(config, selected_ids):
    gmail_addr   = config["gmail_address"].strip()
    app_pass     = config["app_password"].strip()
    save_folder  = config["save_folder"].strip()

    os.makedirs(save_folder, exist_ok=True)

    by_uid = {}
    for item in browse["pdfs"]:
        if item["id"] in selected_ids:
            by_uid.setdefault(item["uid"], []).append(item)

    try:
        server = make_server(gmail_addr, app_pass)
    except Exception as e:
        return {"ok": False, "error": str(e), "saved": [], "failed": []}

    server.select_folder("INBOX", readonly=True)
    saved_files, failed_files = [], []

    for uid, items in by_uid.items():
        try:
            keys = [f"BODY.PEEK[{it['part']}]" for it in items]
            resp = server.fetch([uid], keys)
            data = resp.get(uid, {})
            for it in items:
                raw = data.get(f"BODY[{it['part']}]".encode())
                if raw is None:
                    for k, v in data.items():
                        if k.startswith(b"BODY[") and k.endswith(b"]") and v:
                            raw = v
                            break
                try:
                    pdf_data = _decode_part_body(raw, it["encoding"])
                    _validate_download(pdf_data, it["filename"])
                    kind, save_path = _resolve_save_path(
                        save_folder, it["filename"], it.get("expected", 0))
                    if kind == "exists":
                        saved_files.append(it["filename"] + "  (already on disk)")
                        continue
                    with open(save_path, "wb") as f:
                        f.write(pdf_data)
                    saved_files.append(os.path.basename(save_path))
                except Exception as e:
                    failed_files.append({"file": it["filename"], "error": str(e)})
        except Exception as e:
            for it in items:
                failed_files.append({"file": it["filename"], "error": str(e)})

    try: server.logout()
    except Exception: pass
    return {"ok": True, "saved": saved_files, "failed": failed_files}


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", has_pypdf=HAS_PYPDF)


@app.route("/run", methods=["POST"])
def run():
    if job["running"]:
        return jsonify({"error": "Already running"}), 400
    with _lock:
        if job["running"]:
            return jsonify({"error": "Already running"}), 400
        threading.Thread(target=run_targeted, args=(request.json,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/status")
def status():
    return jsonify({
        "running": job["running"], "done": job["done"],
        "logs": job["logs"][-200:], "rows": job["rows"], "summary": job["summary"]
    })

@app.route("/download")
def download():
    p = job.get("excel_path", "")
    if not p or not os.path.exists(p):
        return "Run an extraction first.", 404
    return send_file(p, as_attachment=True, download_name=os.path.basename(p))


@app.route("/browse", methods=["POST"])
def browse_start():
    if browse["running"]:
        return jsonify({"error": "Browse already running"}), 400
    threading.Thread(target=run_browse, args=(request.json,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/browse_status")
def browse_status():
    return jsonify({
        "running": browse["running"], "done": browse["done"],
        "pdfs": browse["pdfs"], "error": browse["error"]
    })

@app.route("/save_selected", methods=["POST"])
def save_sel():
    data = request.json
    config       = data.get("config", {})
    selected_ids = set(data.get("selected", []))
    if not selected_ids:
        return jsonify({"error": "Nothing selected"}), 400
    result = save_selected(config, selected_ids)
    return jsonify(result)


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)